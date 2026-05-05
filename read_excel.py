import openpyxl

wb = openpyxl.load_workbook('旅游费用明细.xlsx')
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f'=== Sheet: {sheet_name} ===')
    for row in ws.iter_rows(values_only=False):
        vals = [str(cell.value) if cell.value is not None else '' for cell in row]
        print(' | '.join(vals))
    print()
